from openpyxl import Workbook
# import random


class Record:
    def __init__(self, file_name):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.file_name = file_name
        self.current_row = 0

    def save(self):
        self.wb.save(self.file_name)

    def write_layer(self, layer):
        self.current_row += 1
        self.ws.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=16)
        self.ws.cell(row=self.current_row, column=1, value=f'Custom Layer: {layer}')

    def write_header(self):
        self.current_row += 1
        self.ws.merge_cells(start_row=self.current_row, start_column=2, end_row=self.current_row, end_column=4)
        self.ws.cell(row=self.current_row, column=2, value='F1')
        self.ws.merge_cells(start_row=self.current_row, start_column=5, end_row=self.current_row, end_column=7)
        self.ws.cell(row=self.current_row, column=5, value='Accuracy')
        self.ws.merge_cells(start_row=self.current_row, start_column=8, end_row=self.current_row, end_column=10)
        self.ws.cell(row=self.current_row, column=8, value='Precision')
        self.ws.merge_cells(start_row=self.current_row, start_column=11, end_row=self.current_row, end_column=13)
        self.ws.cell(row=self.current_row, column=11, value='Recall')
        self.ws.merge_cells(start_row=self.current_row, start_column=14, end_row=self.current_row, end_column=16)
        self.ws.cell(row=self.current_row, column=14, value='Loss')
        self.current_row += 1
        self.ws.cell(row=self.current_row, column=2, value='Training')
        self.ws.cell(row=self.current_row, column=3, value='Validation')
        self.ws.cell(row=self.current_row, column=4, value='Testing')
        self.ws.cell(row=self.current_row, column=5, value='Training')
        self.ws.cell(row=self.current_row, column=6, value='Validation')
        self.ws.cell(row=self.current_row, column=7, value='Testing')
        self.ws.cell(row=self.current_row, column=8, value='Training')
        self.ws.cell(row=self.current_row, column=9, value='Validation')
        self.ws.cell(row=self.current_row, column=10, value='Testing')
        self.ws.cell(row=self.current_row, column=11, value='Training')
        self.ws.cell(row=self.current_row, column=12, value='Validation')
        self.ws.cell(row=self.current_row, column=13, value='Testing')
        self.ws.cell(row=self.current_row, column=14, value='Training')
        self.ws.cell(row=self.current_row, column=15, value='Validation')
        self.ws.cell(row=self.current_row, column=16, value='Testing')

    def write_base_model(self, base_model):
        self.current_row += 1
        self.ws.cell(row=self.current_row, column=1, value=str(base_model))

    def write_values(self, title, f1, accuracy, precision, recall, loss):
        self.current_row += 1
        column = 1
        self.ws.cell(row=self.current_row, column=column, value=title)
        column += 1
        for i in range(3):
            self.ws.cell(row=self.current_row, column=column, value=f1[i])
            column += 1
        for i in range(3):
            self.ws.cell(row=self.current_row, column=column, value=accuracy[i])
            column += 1
        for i in range(3):
            self.ws.cell(row=self.current_row, column=column, value=precision[i])
            column += 1
        for i in range(3):
            self.ws.cell(row=self.current_row, column=column, value=recall[i])
            column += 1
        for i in range(3):
            self.ws.cell(row=self.current_row, column=column, value=loss[i])
            column += 1

    def write_learn_row_title(self, activation_function):
        self.current_row += 1
        self.ws.cell(row=self.current_row, column=1, value='learning rate')
        self.ws.cell(row=self.current_row, column=2, value=activation_function)


'''
    def test(self):
        self.write_layer('custom layer 2')
        self.write_header()
        self.write_base_model('vgg16')
        activation = ('relu', 'sigmoid', 'tanh')
        for activation in activation:
            f1 = tuple(random.uniform(0, 1) for _ in range(3))
            accuracy = tuple(random.uniform(0, 1) for _ in range(3))
            precision = tuple(random.uniform(0, 1) for _ in range(3))
            recall = tuple(random.uniform(0, 1) for _ in range(3))
            loss = tuple(random.uniform(0, 1) for _ in range(3))
            print(' activation is: ' + activation)
            self.write_values(activation, f1, accuracy, precision, recall, loss)
        best_activation = 'relu'
        self.write_learn_row_title(best_activation)
        learning_rate = [0.001, 0.01, 0.1]
        for lr in learning_rate:
            self.write_values(lr, f1, accuracy, precision, recall, loss)
        self.save()

    # the rest of the code stays the same


test_it = Record('test.xlsx')
test_it.test()
'''
